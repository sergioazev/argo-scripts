import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

FONT = "Arial"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Checklist"

HEADERS = ["Categoria", "Item / Entregável", "Destinatário", "Responsável",
           "Status", "Prazo", "Exigência / Cláusula", "Specs técnicas",
           "Pasta / Local do arquivo", "Observações"]
COL_W = [16, 42, 20, 18, 15, 12, 26, 36, 30, 30]

CAT_COLORS = {
    "DCP": "C9DAF8",
    "MASTERS HD/4K": "D9EAD3",
    "ARQUIVOS DE SOM": "D9D2E9",
    "LEGENDAS E ACESSIBILIDADE": "FCE5CD",
    "DOCUMENTOS": "CCCCCC",
    "MATERIAL PROMOCIONAL": "D0E0E3",
    "DEPÓSITO LEGAL (CINEMATECA/ANCINE)": "F4CCCC",
    "VOD / STREAMING": "C9DAF8",
    "BROADCAST": "EAD1DC",
}

# (categoria, item, exigencia, specs, pasta)
ROWS = [
("DCP", "DCP Brasil — base limpa (sem legendas)", "Contrato distribuidora nacional — preencher cláusula", "Ver skill dcp-validator — SMPTE/Interop, ISDCF naming, reel ≤22min", "/ENTREGAS/[PROJETO]/DCP/BRASIL/"),
("DCP", "DCP Brasil — com legendas em português (se aplicável)", "Contrato distribuidora nacional — preencher cláusula", "Idem acima + legendas queimadas", "/ENTREGAS/[PROJETO]/DCP/BRASIL/"),
("DCP", "DCP Internacional — com legendas em inglês queimadas", "Contrato agente de vendas internacional — preencher cláusula", "Ver skill dcp-validator", "/ENTREGAS/[PROJETO]/DCP/INTERNACIONAL/"),
("DCP", "DCP Internacional — base limpa (sem legendas)", "Contrato agente de vendas internacional — preencher cláusula", "Ver skill dcp-validator", "/ENTREGAS/[PROJETO]/DCP/INTERNACIONAL/"),
("DCP", "DCP outro território/idioma — especificar", "preencher", "preencher", "/ENTREGAS/[PROJETO]/DCP/[TERRITÓRIO]/"),

("MASTERS HD/4K", "Master ProRes 422HQ — base limpa, áudio 5.1", "preencher", "preencher resolução/fps/codec", "/ENTREGAS/[PROJETO]/MASTERS/"),
("MASTERS HD/4K", "Master ProRes 422HQ — com legendas queimadas, áudio 2.0", "preencher", "preencher", "/ENTREGAS/[PROJETO]/MASTERS/"),
("MASTERS HD/4K", "Master 4K DPX/OpenEXR — matriz de preservação (cinema)", "Depósito Legal Cinemateca — ver skill cinemateca-brasileira-deposito", "DCI 24fps, 16bits — ver skill", "/ENTREGAS/[PROJETO]/MASTERS/4K/"),
("MASTERS HD/4K", "Master 4K HDR10 — streaming/VOD", "Plataforma de streaming — preencher", "Ver skill netflix-delivery-specs", "/ENTREGAS/[PROJETO]/MASTERS/HDR10/"),
("MASTERS HD/4K", "Master SDR — trim pass", "preencher", "preencher", "/ENTREGAS/[PROJETO]/MASTERS/SDR/"),

("ARQUIVOS DE SOM", "DME (Dialogue, Music & Effects)", "preencher", "preencher formato/sample rate", "/ENTREGAS/[PROJETO]/SOM/"),
("ARQUIVOS DE SOM", "Final Mix 5.1", "preencher", "preencher", "/ENTREGAS/[PROJETO]/SOM/"),
("ARQUIVOS DE SOM", "Final Mix Stereo 2.0", "preencher", "preencher", "/ENTREGAS/[PROJETO]/SOM/"),
("ARQUIVOS DE SOM", "M&E internacional (versão sem diálogos)", "Contrato agente internacional — preencher", "preencher", "/ENTREGAS/[PROJETO]/SOM/"),
("ARQUIVOS DE SOM", "Trilhas sonoras finais (stems)", "preencher", "preencher", "/ENTREGAS/[PROJETO]/SOM/"),

("LEGENDAS E ACESSIBILIDADE", "Legendas .srt — português", "preencher", "23,98/24/25fps — confirmar", "/ENTREGAS/[PROJETO]/LEGENDAS/"),
("LEGENDAS E ACESSIBILIDADE", "Legendas .srt — inglês", "preencher", "preencher", "/ENTREGAS/[PROJETO]/LEGENDAS/"),
("LEGENDAS E ACESSIBILIDADE", "Legendas .srt — outro idioma (especificar)", "preencher", "preencher", "/ENTREGAS/[PROJETO]/LEGENDAS/"),
("LEGENDAS E ACESSIBILIDADE", "Legendagem descritiva", "ANCINE IN 116/2014 — obrigatório p/ exibição comercial BR", "preencher", "/ENTREGAS/[PROJETO]/ACESSIBILIDADE/"),
("LEGENDAS E ACESSIBILIDADE", "Audiodescrição", "ANCINE IN 116/2014 — obrigatório p/ exibição comercial BR", "preencher", "/ENTREGAS/[PROJETO]/ACESSIBILIDADE/"),
("LEGENDAS E ACESSIBILIDADE", "LIBRAS (Língua Brasileira de Sinais)", "ANCINE IN 116/2014 — obrigatório p/ exibição comercial BR", "preencher", "/ENTREGAS/[PROJETO]/ACESSIBILIDADE/"),

("DOCUMENTOS", "Dialog List", "preencher", "—", "/ENTREGAS/[PROJETO]/DOCUMENTOS/"),
("DOCUMENTOS", "Music Cue Sheet", "preencher", "—", "/ENTREGAS/[PROJETO]/DOCUMENTOS/"),
("DOCUMENTOS", "Chain of Title", "preencher", "—", "/ENTREGAS/[PROJETO]/DOCUMENTOS/"),
("DOCUMENTOS", "Contractual Obligations (exhibit assinado)", "preencher", "—", "/ENTREGAS/[PROJETO]/DOCUMENTOS/"),
("DOCUMENTOS", "Ficha Técnica e Sinopse completa", "preencher", "—", "/ENTREGAS/[PROJETO]/DOCUMENTOS/"),
("DOCUMENTOS", "Certificado de Origem / Film Certificates", "preencher", "—", "/ENTREGAS/[PROJETO]/DOCUMENTOS/"),
("DOCUMENTOS", "Financing Plan", "preencher", "—", "/ENTREGAS/[PROJETO]/DOCUMENTOS/"),

("MATERIAL PROMOCIONAL", "Poster em alta resolução (layers editáveis)", "preencher", "300dpi mín. — formatos .psd/.ai/.indd", "/ENTREGAS/[PROJETO]/PROMO/"),
("MATERIAL PROMOCIONAL", "Stills (10-15 fotos, mín. 300dpi)", "preencher", "preencher", "/ENTREGAS/[PROJETO]/PROMO/"),
("MATERIAL PROMOCIONAL", "Sinopse curta e longa (PT/EN)", "preencher", "logline 2-3 linhas / sinopse 5-7 linhas", "/ENTREGAS/[PROJETO]/PROMO/"),
("MATERIAL PROMOCIONAL", "Trailer", "preencher", "preencher", "/ENTREGAS/[PROJETO]/PROMO/"),
("MATERIAL PROMOCIONAL", "Press kit / Bio e filmografia do diretor", "preencher", "—", "/ENTREGAS/[PROJETO]/PROMO/"),
("MATERIAL PROMOCIONAL", "Billing block editável + logotipos vetorizados", "preencher", "fonte anexada + vetor", "/ENTREGAS/[PROJETO]/PROMO/"),

("DEPÓSITO LEGAL (CINEMATECA/ANCINE)", "Matriz Cinema (DPX/EXR/TIFF, 24fps DCI)", "Lei do Audiovisual 8685 / MP 2.228 — ver skill cinemateca-brasileira-deposito", "ver skill", "/ENTREGAS/[PROJETO]/DEPOSITO_LEGAL/"),
("DEPÓSITO LEGAL (CINEMATECA/ANCINE)", "Matriz TV (MKV + FFV1v3, 23.976/29.97/59.94)", "idem", "ver skill", "/ENTREGAS/[PROJETO]/DEPOSITO_LEGAL/"),
("DEPÓSITO LEGAL (CINEMATECA/ANCINE)", "Ficha Técnica Resumida", "idem", "formulário oficial Cinemateca", "/ENTREGAS/[PROJETO]/DEPOSITO_LEGAL/"),
("DEPÓSITO LEGAL (CINEMATECA/ANCINE)", "Cadastro do Depositante", "idem", "formulário oficial Cinemateca", "/ENTREGAS/[PROJETO]/DEPOSITO_LEGAL/"),
("DEPÓSITO LEGAL (CINEMATECA/ANCINE)", "Suporte físico (LTO-9 ou CRU DX115)", "idem", "ver skill", "/ENTREGAS/[PROJETO]/DEPOSITO_LEGAL/"),
("DEPÓSITO LEGAL (CINEMATECA/ANCINE)", "Laudo técnico ANCINE", "idem", "ver tabela de laudos Ancine 2026", "/ENTREGAS/[PROJETO]/DEPOSITO_LEGAL/"),

("VOD / STREAMING", "Master IMF (App2 Extended) — conforme plataforma", "Contrato plataforma — preencher", "Ver skill netflix-delivery-specs", "/ENTREGAS/[PROJETO]/VOD/"),
("VOD / STREAMING", "QC de loudness (LUFS) e True Peak", "preencher", "Ver skill ffmpeg-qc-commands", "/ENTREGAS/[PROJETO]/VOD/"),
("VOD / STREAMING", "Subtitles IMSC1 / timed text", "preencher", "preencher", "/ENTREGAS/[PROJETO]/VOD/"),
("VOD / STREAMING", "Metadados de entrega (nomenclatura ISDCF + código do estúdio)", "preencher", "preencher", "/ENTREGAS/[PROJETO]/VOD/"),

("BROADCAST", "Master conforme especificação do canal — preencher", "Contrato canal — preencher", "preencher", "/ENTREGAS/[PROJETO]/BROADCAST/"),
("BROADCAST", "QC técnico (broadcast range, black/freeze/silence)", "preencher", "Ver skill ffmpeg-qc-commands", "/ENTREGAS/[PROJETO]/BROADCAST/"),
]

STATUS_OPTIONS = ["Não iniciado", "Em produção", "Entregue", "Aprovado", "N/A"]

thin = Side(style="thin", color="BFBFBF")
border = Border(top=thin, bottom=thin, left=thin, right=thin)

# Title
ws.merge_cells("A1:J1")
c = ws["A1"]
c.value = "CHECKLIST DE DELIVERIES — [NOME DO PROJETO]"
c.font = Font(name=FONT, bold=True, size=16)
c.alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells("A2:J2")
c = ws["A2"]
c.value = "Modelo padrão Argonautas — copiar esta aba para cada novo projeto e preencher Destinatário / Responsável / Prazo / Pasta. Apagar linhas que não se aplicam."
c.font = Font(name=FONT, italic=True, size=10, color="777777")

HEADER_ROW = 4
for j, h in enumerate(HEADERS, start=1):
    cell = ws.cell(HEADER_ROW, j, h)
    cell.font = Font(name=FONT, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="434343")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

start_row = HEADER_ROW + 1
r = start_row
for cat, item, exig, specs, pasta in ROWS:
    fill = PatternFill("solid", fgColor=CAT_COLORS.get(cat, "FFFFFF"))
    vals = [cat, item, "", "", "Não iniciado", "", exig, specs, pasta, ""]
    for j, v in enumerate(vals, start=1):
        cell = ws.cell(r, j, v)
        cell.font = Font(name=FONT, size=10, bold=(j == 1))
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=(j in (2, 7, 8, 9, 10)))
        cell.border = border
    r += 1
end_row = r - 1

for j, w in enumerate(COL_W, start=1):
    ws.column_dimensions[get_column_letter(j)].width = w

ws.freeze_panes = f"A{start_row}"

# Excel Table (filter/sort) without built-in banding so category fills stay visible
tbl = Table(displayName="Deliveries", ref=f"A{HEADER_ROW}:J{end_row}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                     showFirstColumn=False, showLastColumn=False)
ws.add_table(tbl)

# Status dropdown
dv = DataValidation(type="list", formula1='"' + ",".join(STATUS_OPTIONS) + '"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"E{start_row}:E{end_row}")

# Conditional formatting on Status
status_fmt = {
    "Entregue": ("C6EFCE", "006100"),
    "Aprovado": ("A9D08E", "1E4620"),
    "Em produção": ("FFEB9C", "9C6500"),
    "Não iniciado": ("FFC7CE", "9C0006"),
    "N/A": ("D9D9D9", "7F7F7F"),
}
rng = f"E{start_row}:E{end_row}"
for status, (bg, fg) in status_fmt.items():
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=[f'"{status}"'],
                   fill=PatternFill("solid", fgColor=bg), font=Font(name=FONT, color=fg, bold=True))
    )

# --- Resumo sheet ---
ws2 = wb.create_sheet("Resumo")
ws2.merge_cells("A1:C1")
ws2["A1"] = "RESUMO — STATUS DE DELIVERIES"
ws2["A1"].font = Font(name=FONT, bold=True, size=14)

hdr2 = ["Status", "Qtd", "% do total"]
for j, h in enumerate(hdr2, start=1):
    cell = ws2.cell(3, j, h)
    cell.font = Font(name=FONT, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="434343")

row2 = 4
status_rows = {}
for status in STATUS_OPTIONS:
    ws2.cell(row2, 1, status).font = Font(name=FONT, size=10)
    ws2.cell(row2, 2, f'=COUNTIF(Checklist!E{start_row}:E{end_row},"{status}")').font = Font(name=FONT, size=10)
    status_rows[status] = row2
    row2 += 1

total_row = row2
ws2.cell(total_row, 1, "Total de itens").font = Font(name=FONT, bold=True)
ws2.cell(total_row, 2, f'=COUNTA(Checklist!B{start_row}:B{end_row})').font = Font(name=FONT, bold=True)

for status, rr2 in status_rows.items():
    ws2.cell(rr2, 3, f"=B{rr2}/B{total_row}")
    ws2.cell(rr2, 3).number_format = "0%"

pct_row = total_row + 2
ws2.cell(pct_row, 1, "% concluído (Entregue + Aprovado)").font = Font(name=FONT, bold=True)
ws2.cell(pct_row, 2, f"=(B{status_rows['Entregue']}+B{status_rows['Aprovado']})/B{total_row}").number_format = "0%"
ws2.cell(pct_row, 2).font = Font(name=FONT, bold=True)

# Breakdown by category
cat_start = pct_row + 3
ws2.cell(cat_start - 1, 1, "Por categoria").font = Font(name=FONT, bold=True, size=12)
hdr3 = ["Categoria", "Itens", "Concluídos", "% concluído"]
for j, h in enumerate(hdr3, start=1):
    cell = ws2.cell(cat_start, j, h)
    cell.font = Font(name=FONT, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="666666")

cats_unique = list(dict.fromkeys(c for c, *_ in ROWS))
rr = cat_start + 1
for cat in cats_unique:
    ws2.cell(rr, 1, cat).font = Font(name=FONT, size=10)
    ws2.cell(rr, 2, f'=COUNTIF(Checklist!A{start_row}:A{end_row},"{cat}")').font = Font(name=FONT, size=10)
    ws2.cell(rr, 3, f'=COUNTIFS(Checklist!A{start_row}:A{end_row},"{cat}",Checklist!E{start_row}:E{end_row},"Entregue")'
                     f'+COUNTIFS(Checklist!A{start_row}:A{end_row},"{cat}",Checklist!E{start_row}:E{end_row},"Aprovado")').font = Font(name=FONT, size=10)
    ws2.cell(rr, 4, f"=C{rr}/B{rr}").number_format = "0%"
    rr += 1

for col, w in zip("ABCD", [38, 10, 12, 12]):
    ws2.column_dimensions[col].width = w

wb.save("/sessions/youthful-affectionate-galileo/mnt/outputs/checklist-deliveries-argonautas.xlsx")
print("saved", start_row, end_row)
